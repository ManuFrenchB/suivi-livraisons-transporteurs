#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rapport_mail.py - Genere le rapport de suivi et l'envoie par mail (French Bloom)

Fusionne monitoring_expeditions.csv (SharePoint) + suivi_resultats.csv (scraper),
produit un Excel priorise (retards en tete) et l'envoie par mail via Graph.

Variables d'environnement :
  MS_TENANT_ID / MS_CLIENT_ID / MS_CLIENT_SECRET   (alias AZURE_* acceptes)
  SEND_MAIL "1" pour envoyer ; RAPPORT_FROM ; RAPPORT_TO ; RAPPORT_CC (option)

Dependances : pip install msal requests openpyxl
"""

import os
import csv
import sys
import base64
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GRAPH = "https://graph.microsoft.com/v1.0"

# Alias de colonnes tolerees (le vrai fichier utilise numero_commande / pays)
COMMANDE = ("numero_commande", "fnumero_commande")
LIEN_BONX = ("lien_bonx",)
TRANSPORTEUR = ("transporteur",)
NUMERO_SUIVI = ("numero_suivi",)
DATE_SOUH = ("date_livraison_souhaitee",)
DATE_SOUH_P1 = ("date_souhaitee_plus_1j_ouvre",)
DATE_REELLE = ("date_livraison_reelle",)
PAYS = ("pays", "rpays")
VILLE = ("ville",)
PIPELINE = ("pipeline",)

DATE_FORMATS = ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y", "%Y/%m/%d")

ST_INCIDENT = "Incident"
ST_RETARD = "En retard"
ST_AVERIF = "À vérifier"
ST_ENCOURS = "En cours"
ST_LIVRE_RET = "Livré en retard"
ST_LIVRE = "Livré"
ST_INDET = "Indéterminé"

PRIORITE = {ST_INCIDENT: 0, ST_RETARD: 1, ST_AVERIF: 2, ST_ENCOURS: 3,
            ST_LIVRE_RET: 4, ST_LIVRE: 5, ST_INDET: 6}
ORDRE = [ST_INCIDENT, ST_RETARD, ST_AVERIF, ST_ENCOURS, ST_LIVRE_RET, ST_LIVRE, ST_INDET]


def val(row, keys):
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return ""


def env_any(*names, required=False, default=None):
    for n in names:
        v = os.environ.get(n)
        if v:
            return v
    if required:
        raise KeyError("Variable manquante : " + " / ".join(names))
    return default


def parse_date(v):
    if not v:
        return None
    s = str(v).strip().split(" ")[0].split("T")[0]
    if not s or s.lower() in ("none", "nan", "null"):
        return None
    for f in DATE_FORMATS:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def read_csv(path):
    if not os.path.exists(path):
        return []
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                c = {",": sample.count(","), ";": sample.count(";"), "\t": sample.count("\t")}
                delim = max(c, key=c.get)
                return [{(k or "").strip(): v for k, v in r.items()}
                        for r in csv.DictReader(f, delimiter=delim)]
        except UnicodeDecodeError:
            continue
    return []


def normaliser_scraper(raw):
    t = (raw or "").strip().lower()
    if not t:
        return ST_AVERIF
    if "incident" in t:
        return ST_INCIDENT
    if "retard" in t:
        return ST_RETARD
    if t.startswith("livr"):
        return ST_LIVRE
    if "cours" in t or "transit" in t:
        return ST_ENCOURS
    return ST_AVERIF


def statut_final(row, scrap, auj):
    d_souh = parse_date(val(row, DATE_SOUH))
    d_p1 = parse_date(val(row, DATE_SOUH_P1)) or d_souh
    d_reel = parse_date(val(row, DATE_REELLE))
    if d_reel is not None:
        ref = d_p1 or d_souh
        if ref and d_reel > ref:
            return ST_LIVRE_RET, (d_reel - ref).days
        return ST_LIVRE, 0
    if scrap:
        st = normaliser_scraper(scrap.get("statut_normalise", ""))
        try:
            jr = int(scrap.get("jours_retard", "0") or 0)
        except ValueError:
            jr = 0
        ref = d_p1 or d_souh
        if st == ST_RETARD and jr == 0 and ref and auj > ref:
            jr = (auj - ref).days
        return st, jr
    ref = d_p1 or d_souh
    if ref and auj > ref:
        return ST_RETARD, (auj - ref).days
    if d_souh and auj >= d_souh:
        return ST_AVERIF, 0
    return ST_ENCOURS, 0


FILLS = {
    ST_INCIDENT: PatternFill("solid", fgColor="EA9999"),
    ST_RETARD: PatternFill("solid", fgColor="F4CCCC"),
    ST_AVERIF: PatternFill("solid", fgColor="FCE5CD"),
    ST_ENCOURS: PatternFill("solid", fgColor="FFFFFF"),
    ST_LIVRE_RET: PatternFill("solid", fgColor="FFF2CC"),
    ST_LIVRE: PatternFill("solid", fgColor="D9EAD3"),
    ST_INDET: PatternFill("solid", fgColor="EFEFEF"),
}
HF = PatternFill("solid", fgColor="1F3864")
HFONT = Font(color="FFFFFF", bold=True)
LFONT = Font(color="1155CC", underline="single")
TH = Side(style="thin", color="D0D0D0")
BORD = Border(left=TH, right=TH, top=TH, bottom=TH)
ENT = ["N° commande", "Statut", "Jours de retard", "Transporteur", "Statut transporteur",
       "N° / lien de suivi", "Pays", "Ville", "Date souhaitée", "Date souhaitée +1j",
       "Date réelle", "Pipeline", "Fiche Bonx"]


def _rang(m):
    livree = 1 if val(m, DATE_REELLE) else 0
    suivi = val(m, NUMERO_SUIVI).lower()
    url = 1 if suivi.startswith(("http://", "https://")) else 0
    return (livree, url)


def build_rows(auj):
    master = read_csv("monitoring_expeditions.csv")
    scrap_list = read_csv("suivi_resultats.csv")
    scrap = {(r.get("numero_commande") or "").strip(): r for r in scrap_list}

    dedup = {}
    for m in master:
        cmd = val(m, COMMANDE)
        if cmd and (cmd not in dedup or _rang(m) >= _rang(dedup[cmd])):
            dedup[cmd] = m
    master = list(dedup.values())

    rows = []
    for m in master:
        cmd = val(m, COMMANDE)
        s = scrap.get(cmd)
        statut, jr = statut_final(m, s, auj)
        rows.append({
            "cmd": cmd, "statut": statut, "jr": jr,
            "transp": val(m, TRANSPORTEUR),
            "statut_transp": (s or {}).get("statut_brut", "").strip() if s else "",
            "suivi": val(m, NUMERO_SUIVI),
            "pays": val(m, PAYS),
            "ville": val(m, VILLE),
            "d_souh": val(m, DATE_SOUH),
            "d_p1": val(m, DATE_SOUH_P1),
            "d_reel": val(m, DATE_REELLE),
            "pipeline": val(m, PIPELINE),
            "lien_bonx": val(m, LIEN_BONX),
        })
    rows.sort(key=lambda r: (PRIORITE.get(r["statut"], 9), -r["jr"]))
    return rows


def generer_xlsx(rows, auj, out):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Synthèse"
    cpt = {}
    for r in rows:
        cpt[r["statut"]] = cpt.get(r["statut"], 0) + 1
    ws0["A1"] = "Suivi des livraisons - French Bloom"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"] = "Généré le " + auj.strftime("%d/%m/%Y")
    ws0["A2"].font = Font(italic=True, color="666666")
    ws0["A4"] = "Statut"
    ws0["B4"] = "Nombre"
    ws0["A4"].font = Font(bold=True)
    ws0["B4"].font = Font(bold=True)
    ln = 5
    for s in ORDRE:
        if cpt.get(s):
            ws0.cell(ln, 1, s).fill = FILLS.get(s, PatternFill())
            ws0.cell(ln, 2, cpt[s])
            ln += 1
    ws0.cell(ln, 1, "TOTAL").font = Font(bold=True)
    ws0.cell(ln, 2, len(rows)).font = Font(bold=True)
    ws0.column_dimensions["A"].width = 20
    ws0.column_dimensions["B"].width = 12

    ws = wb.create_sheet("Détail")
    for c, t in enumerate(ENT, 1):
        cell = ws.cell(1, c, t)
        cell.fill = HF
        cell.font = HFONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORD
    ws.freeze_panes = "A2"
    for i, r in enumerate(rows, 2):
        suivi = r["suivi"]
        url = suivi if suivi.lower().startswith(("http://", "https://")) else None
        vals = [r["cmd"], r["statut"], r["jr"] if r["jr"] else "", r["transp"],
                r["statut_transp"], suivi, r["pays"], r["ville"], r["d_souh"],
                r["d_p1"], r["d_reel"], r["pipeline"], "Ouvrir" if r["lien_bonx"] else ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.border = BORD
            cell.alignment = Alignment(vertical="center")
        fill = FILLS.get(r["statut"])
        if fill:
            for c in range(1, len(ENT) + 1):
                ws.cell(i, c).fill = fill
        if url:
            ws.cell(i, 6).hyperlink = url
            ws.cell(i, 6).font = LFONT
        if r["lien_bonx"]:
            ws.cell(i, len(ENT)).hyperlink = r["lien_bonx"]
            ws.cell(i, len(ENT)).font = LFONT
    for c, w in enumerate([15, 15, 13, 13, 22, 26, 10, 15, 14, 16, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.auto_filter.ref = "A1:" + get_column_letter(len(ENT)) + str(len(rows) + 1)
    wb.save(out)
    return cpt


def corps_html(rows, cpt, auj):
    urgents = [r for r in rows if r["statut"] in (ST_RETARD, ST_INCIDENT)]
    lignes = ""
    for r in urgents[:40]:
        lignes += ("<tr><td>" + r["cmd"] + "</td><td>" + r["statut"] + "</td>"
                   "<td style='text-align:center'>" + (str(r["jr"]) if r["jr"] else "") + "</td>"
                   "<td>" + r["transp"] + "</td><td>" + r["ville"] + " (" + r["pays"] + ")</td></tr>")
    synth = " &middot; ".join(s + " : <b>" + str(cpt[s]) + "</b>" for s in ORDRE if cpt.get(s))
    if urgents:
        bloc = ("<p><b>" + str(len(urgents)) + " commande(s) à traiter en priorité</b> (retard / incident) :</p>"
                "<table border='1' cellpadding='6' cellspacing='0' style='border-collapse:collapse;font-family:Arial;font-size:13px'>"
                "<tr style='background:#1F3864;color:#fff'><th>Commande</th><th>Statut</th>"
                "<th>Jours retard</th><th>Transporteur</th><th>Destination</th></tr>" + lignes + "</table>")
    else:
        bloc = "<p style='color:#38761d'><b>Aucune livraison en retard ni en incident.</b></p>"
    return ("<div style='font-family:Arial;font-size:14px;color:#222'>"
            "<h2 style='color:#1F3864'>Suivi des livraisons - " + auj.strftime("%d/%m/%Y") + "</h2>"
            "<p>" + synth + "</p>" + bloc +
            "<p style='color:#666;font-size:12px'>Détail complet en piece jointe (Excel). "
            "Rapport automatique French Bloom.</p></div>")


def get_token():
    import msal
    tenant = env_any("MS_TENANT_ID", "AZURE_TENANT_ID", required=True)
    client_id = env_any("MS_CLIENT_ID", "AZURE_CLIENT_ID", required=True)
    secret = env_any("MS_CLIENT_SECRET", "AZURE_CLIENT_SECRET", required=True)
    app = msal.ConfidentialClientApplication(
        client_id=client_id, client_credential=secret,
        authority="https://login.microsoftonline.com/" + tenant)
    res = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in res:
        raise RuntimeError("Auth Graph echouee : " + str(res.get("error_description", res)))
    return res["access_token"]


def envoyer_mail(sujet, html, piece_jointe):
    import requests
    sender = os.environ["RAPPORT_FROM"]
    to = [a.strip() for a in os.environ.get("RAPPORT_TO", "").split(",") if a.strip()]
    cc = [a.strip() for a in os.environ.get("RAPPORT_CC", "").split(",") if a.strip()]
    if not to:
        print("RAPPORT_TO vide : envoi ignore.")
        return
    with open(piece_jointe, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode()
    message = {
        "message": {
            "subject": sujet,
            "body": {"contentType": "HTML", "content": html},
            "toRecipients": [{"emailAddress": {"address": a}} for a in to],
            "ccRecipients": [{"emailAddress": {"address": a}} for a in cc],
            "attachments": [{
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": os.path.basename(piece_jointe),
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentBytes": content_b64,
            }],
        },
        "saveToSentItems": True,
    }
    token = get_token()
    r = requests.post(GRAPH + "/users/" + sender + "/sendMail",
                      headers={"Authorization": "Bearer " + token, "Content-Type": "application/json"},
                      json=message, timeout=60)
    r.raise_for_status()
    print("Mail envoye a " + ", ".join(to))


def main():
    auj = date.today()
    rows = build_rows(auj)
    if not rows:
        print("Aucune donnee (monitoring_expeditions.csv manquant ou vide).")
        return
    out = "suivi_livraisons_" + auj.isoformat() + ".xlsx"
    cpt = generer_xlsx(rows, auj, out)
    print(str(len(rows)) + " commandes -> " + out)
    for s in ORDRE:
        if cpt.get(s):
            print("  " + s + ": " + str(cpt[s]))
    nb_urgent = cpt.get(ST_RETARD, 0) + cpt.get(ST_INCIDENT, 0)
    sujet = "[Suivi livraisons] " + auj.strftime("%d/%m/%Y") + " - " + str(nb_urgent) + " a traiter"
    html = corps_html(rows, cpt, auj)
    if os.environ.get("SEND_MAIL") == "1":
        envoyer_mail(sujet, html, out)
    else:
        with open("apercu_mail_" + auj.isoformat() + ".html", "w", encoding="utf-8") as f:
            f.write(html)
        print("SEND_MAIL != 1 : apercu HTML ecrit, aucun mail envoye.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ERREUR rapport_mail : " + str(e), file=sys.stderr)
        sys.exit(1)
